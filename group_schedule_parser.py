import re
import os
import requests
from bs4 import BeautifulSoup
import openpyxl
import json

weekdays = ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"]


def group_schedule_parser():
    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")

    result = soup.find("div", {"class": "rasspisanie"}). \
        find(string="Институт информационных технологий"). \
        find_parent("div").find_parent("div").find_all("a", {"class": "uk-link-toggle"})

    file_data = {"groups": []}
    different_teachers = []

    for a_tag in result:
        if "ИИТ_4" in a_tag['href']:
            continue

        table = requests.get(a_tag['href'])

        with open("table.xlsx", "wb") as f:
            f.write(table.content)

        book = openpyxl.load_workbook("table.xlsx")
        sheet = book.active
        num_cols = sheet.max_column

        for col in range(1, num_cols):
            group = str(sheet.cell(row=2, column=col).value)
            if re.search(r"[а-яА-Я]{4}\-\d\d\-\d\d", group):
                group_data = {"group": group, "timetable": {}}

                schedule_data = {"odd": {}, "even": {}}

                for d in weekdays:
                    schedule_data["odd"].update({d: {}})
                    schedule_data["even"].update({d: {}})
                    for n in "123456":
                        schedule_data["odd"][d].update({n: {}})
                        schedule_data["even"][d].update({n: {}})
                        for f in "subject", "teacher", "type", "aud":
                            schedule_data["odd"][d][n].update({f: '-'})
                            schedule_data["even"][d][n].update({f: '-'})

                i = 0
                for row in range(4, 76):
                    subject = sheet.cell(row=row, column=col).value
                    sub_type = sheet.cell(row=row, column=col + 1).value
                    teacher = sheet.cell(row=row, column=col + 2).value
                    aud = sheet.cell(row=row, column=col + 3).value

                    if (row - 4) % 12 == 0:
                        if row != 4:
                            i += 1

                    if row % 2 == 0:
                        num = str((((row - 4) % 12) + 2) // 2)
                        col_name = "odd"
                    else:
                        num = str((((row - 4) % 12) + 1) // 2)
                        col_name = "even"

                    if re.search(r"[А-Яа-яёЁ]+", str(subject)):

                        schedule_data[col_name][weekdays[i]][num]["subject"] = str(subject)
                        if teacher:

                            exclusives = ["Гриценко", "Ноовсёлова", "Молчанов"]
                            t_list = []
                            inside = False
                            for ex in exclusives:
                                if ex in str(teacher):
                                    inside = True
                                    break

                            if not inside:
                                ff = re.findall(r"[А-Яа-яЁё]+-?[А-Яа-яЁё]+ +[A-ЯЁ][., ]+[A-ЯЁ][., ]?", str(teacher))

                            else:
                                if exclusives[0] not in different_teachers:
                                    different_teachers.append(exclusives[0])
                                ff = str(teacher).split("\n")

                            for match in ff:
                                match = match.replace("  ", " ")
                                match = match.replace("..", ".")
                                match = match.strip()

                                if match[-1] == ".":
                                    match = match[:-1]

                                if match == "Иоффе Н Е":
                                    match = "Иоффе Н.Е"
                                elif match == "Ануфриев О. С":
                                    match = "Ануфриев О.С"
                                elif match == "Ноовсёлова":
                                    match = "Новосёлова Е.В"
                                elif match == "Молчанова" or match == "Молчановва":
                                    match = "Молчанова И.В"

                                if match[-1] == ".":
                                    match = match[:-1]
                                t_list.append(match)
                                if match not in different_teachers:
                                    different_teachers.append(match)

                            if len(t_list) > 0:
                                schedule_data[col_name][weekdays[i]][num]["teacher"] = "\n".join(t_list)

                        if sub_type:
                            schedule_data[col_name][weekdays[i]][num]["type"] = str(sub_type)
                        if aud:
                            schedule_data[col_name][weekdays[i]][num]["aud"] = str(aud)

                group_data["timetable"] = schedule_data
                file_data["groups"].append(group_data)

    filepath1 = os.path.abspath(os.curdir) + "/tables/groups_schedule.json"
    with open(filepath1, 'w', encoding="utf-8") as file:
        json.dump(file_data, file, indent=4, ensure_ascii=False)

    filepath2 = os.path.abspath(os.curdir) + "/tables/teachers.json"
    with open(filepath2, 'w', encoding="utf-8") as file:
        json.dump({"teachers": different_teachers}, file, indent=4, ensure_ascii=False)
